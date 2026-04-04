#!/usr/bin/env python3
"""
scripts/import_data.py

Seeds the SQLite database from one of two sources:
  1.  A JSON file produced by extract_from_html.py  (default)
  2.  An Excel file (RawData.xlsx) if --excel is passed

Usage
-----
# From JSON (fastest, use after extract_from_html.py)
python scripts/import_data.py

# From Excel directly (requires openpyxl)
python scripts/import_data.py --excel data/RawData.xlsx --sheet "DataEntry"

Options
-------
--json PATH     Path to JSON file  (default: data/records.json)
--excel PATH    Path to Excel file
--sheet NAME    Sheet name in Excel  (default: first sheet)
--clear         Drop all existing records before importing
"""
import argparse
import json
import sys
import os

# Make sure the project root is on sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import create_tables, SessionLocal, Record

# ── Column mapping: JSON/Excel key → ORM field name ──────────
FIELD_MAP = {
    "zone":            "zone",
    "scheme":          "scheme",
    "month":           "month",
    "monthNo":         "month_no",
    "year":            "year",
    "quarter":         "quarter",
    "volProduced":     "vol_produced",
    "revenueWater":    "revenue_water",
    "nrw":             "nrw",
    "pctNRW":          "pct_nrw",
    "chemCost":        "chem_cost",
    "powerCost":       "power_cost",
    "powerKwh":        "power_kwh",
    "fuelCost":        "fuel_cost",
    "staffCosts":      "staff_costs",
    "opCost":          "op_cost",
    "newConnections":  "new_connections",
    "activeCustomers": "active_customers",
    "activePostpaid":  "active_postpaid",
    "activePrepaid":   "active_prepaid",
    "stuckMeters":     "stuck_meters",
    "stuckNew":        "stuck_new",
    "stuckRepaired":   "stuck_repaired",
    "pipeBreakdowns":  "pipe_breakdowns",
    "pumpBreakdowns":  "pump_breakdowns",
    "supplyHours":     "supply_hours",
    "powerFailHours":  "power_fail_hours",
    "cashCollected":   "cash_collected",
    "amtBilled":       "amt_billed",
    "serviceCharge":   "service_charge",
    "meterRental":     "meter_rental",
    "totalSales":      "total_sales",
    "privateDebtors":  "private_debtors",
    "publicDebtors":   "public_debtors",
    "totalDebtors":    "total_debtors",
    "collectionRate":  "collection_rate",
    "popSupplied":     "pop_supplied",
    "popSupplyArea":   "pop_supply_area",
    "permStaff":       "perm_staff",
    "tempStaff":       "temp_staff",
}


def map_row(raw: dict) -> dict:
    """Convert camelCase dict (JSON/Excel) to ORM snake_case dict."""
    mapped = {}
    for src, dst in FIELD_MAP.items():
        if src in raw:
            v = raw[src]
            mapped[dst] = float(v) if isinstance(v, (int, float)) and dst not in ("zone","scheme","month","quarter") else v
    return mapped


def import_from_json(path: str) -> list:
    with open(path) as f:
        return json.load(f)


def import_from_excel(path: str, sheet: str | None = None) -> list:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    print(f"  Read {len(rows)} rows from '{ws.title}'")
    return rows


def seed(raw_rows: list, clear: bool = False) -> tuple[int, int, list]:
    create_tables()
    db = SessionLocal()
    inserted = skipped = 0
    errors = []

    try:
        if clear:
            deleted = db.query(Record).delete()
            db.commit()
            print(f"  Cleared {deleted} existing records")

        for i, raw in enumerate(raw_rows):
            try:
                data = map_row(raw)
                if not all(k in data for k in ("zone", "scheme", "month", "year")):
                    skipped += 1
                    errors.append(f"Row {i}: missing required identity fields")
                    continue

                # Upsert: skip if already present (use --clear to replace all)
                exists = (
                    db.query(Record)
                    .filter(
                        Record.zone   == data["zone"],
                        Record.scheme == data["scheme"],
                        Record.month  == data["month"],
                        Record.year   == int(data["year"]),
                    )
                    .first()
                )
                if exists:
                    skipped += 1
                    continue

                db.add(Record(**data))
                inserted += 1

                if inserted % 100 == 0:
                    db.commit()

            except Exception as e:
                errors.append(f"Row {i}: {e}")
                skipped += 1

        db.commit()
    finally:
        db.close()

    return inserted, skipped, errors


def main():
    parser = argparse.ArgumentParser(description="Seed SRWB SQLite database")
    parser.add_argument("--json",  default="data/records.json", help="Path to JSON file")
    parser.add_argument("--excel", default=None,                help="Path to Excel file")
    parser.add_argument("--sheet", default=None,                help="Excel sheet name")
    parser.add_argument("--clear", action="store_true",         help="Clear existing records first")
    args = parser.parse_args()

    print("SRWB Data Import")
    print("=" * 40)

    if args.excel:
        print(f"  Source : Excel → {args.excel}")
        raw_rows = import_from_excel(args.excel, args.sheet)
    else:
        json_path = args.json
        if not os.path.exists(json_path):
            print(f"ERROR: JSON file not found: {json_path}")
            print("Run scripts/extract_from_html.py first, or pass --excel <path>")
            sys.exit(1)
        print(f"  Source : JSON → {json_path}")
        raw_rows = import_from_json(json_path)

    print(f"  Rows   : {len(raw_rows)}")
    print(f"  Clear  : {'yes' if args.clear else 'no (skip duplicates)'}")
    print()

    inserted, skipped, errors = seed(raw_rows, clear=args.clear)

    print(f"  ✓ Inserted : {inserted}")
    print(f"  ⊘ Skipped  : {skipped}")
    if errors:
        print(f"  ✗ Errors   : {len(errors)}")
        for e in errors[:10]:
            print(f"    {e}")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
